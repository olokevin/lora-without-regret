"""Build xlsx copies of the markdown result tables in this directory.

Each .md source becomes one .xlsx; each markdown table within it becomes one sheet.
Numeric cells are written as numbers (so sorting / formulas work); bold-marker
asterisks and code-backticks in cells are stripped. Bold cells in the source
get Excel bold formatting.
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

HERE = Path(__file__).parent

BOLD_RE = re.compile(r"\*\*(.+?)\*\*")
CODE_RE = re.compile(r"`([^`]*)`")


def split_row(line: str) -> list[str]:
    s = line.strip()
    if s.startswith("|"):
        s = s[1:]
    if s.endswith("|"):
        s = s[:-1]
    return [c.strip() for c in s.split("|")]


def is_separator(cells: list[str]) -> bool:
    return bool(cells) and all(re.fullmatch(r":?-+:?", c or "") for c in cells)


def clean_cell(raw: str) -> tuple[object, bool]:
    """Return (value, is_bold). Converts numeric cells to float/int."""
    text = raw
    bold = False
    m = BOLD_RE.search(text)
    if m:
        bold = True
        text = BOLD_RE.sub(r"\1", text)
    text = CODE_RE.sub(r"\1", text).strip()
    if text in {"", "-"}:
        return (text, bold)
    try:
        if re.fullmatch(r"-?\d+", text):
            return (int(text), bold)
        return (float(text), bold)
    except ValueError:
        return (text, bold)


def parse_tables(md: str) -> list[tuple[str, list[list[str]]]]:
    """Parse markdown into [(heading, rows)]. rows include the header row."""
    lines = md.splitlines()
    tables: list[tuple[str, list[list[str]]]] = []
    current_heading = ""
    i = 0
    while i < len(lines):
        line = lines[i]
        stripped = line.strip()
        if stripped.startswith("#"):
            current_heading = stripped.lstrip("#").strip()
            i += 1
            continue
        if stripped.startswith("|") and i + 1 < len(lines):
            nxt = split_row(lines[i + 1])
            if is_separator(nxt):
                header = split_row(line)
                rows = [header]
                j = i + 2
                while j < len(lines) and lines[j].strip().startswith("|"):
                    rows.append(split_row(lines[j]))
                    j += 1
                tables.append((current_heading, rows))
                i = j
                continue
        i += 1
    return tables


def unique_sheet_name(wb: Workbook, base: str) -> str:
    base = re.sub(r"[\[\]\*\?/\\:]", "", base).strip() or "Sheet"
    base = base[:31]
    if base not in wb.sheetnames:
        return base
    for k in range(2, 100):
        cand = f"{base[:31 - len(str(k)) - 1]} {k}"
        if cand not in wb.sheetnames:
            return cand
    return base


def convert(md_path: Path, xlsx_path: Path) -> int:
    tables = parse_tables(md_path.read_text())
    if not tables:
        return 0
    wb = Workbook()
    wb.remove(wb.active)
    bold_font = Font(bold=True)
    header_font = Font(bold=True)
    for heading, rows in tables:
        header = rows[0]
        body = [r for r in rows[1:] if any((c or "").strip() for c in r)]
        kept_rows = [header] + body
        sheet_base = heading.split(" (")[0] or "Table"
        ws = wb.create_sheet(unique_sheet_name(wb, sheet_base))
        for r_idx, row in enumerate(kept_rows, start=1):
            for c_idx, raw in enumerate(row, start=1):
                value, is_bold = clean_cell(raw)
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = header_font
                elif is_bold:
                    cell.font = bold_font
        rows = kept_rows
        for col_idx, _ in enumerate(rows[0], start=1):
            max_len = max(
                (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, len(rows) + 1)),
                default=10,
            )
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 60)
    wb.save(xlsx_path)
    return len(tables)


def main() -> None:
    for md in sorted(HERE.glob("*.md")):
        if md.name == "claude.md":
            continue
        xlsx = md.with_suffix(".xlsx")
        n = convert(md, xlsx)
        print(f"{md.name} -> {xlsx.name} ({n} table{'s' if n != 1 else ''})")


if __name__ == "__main__":
    main()
